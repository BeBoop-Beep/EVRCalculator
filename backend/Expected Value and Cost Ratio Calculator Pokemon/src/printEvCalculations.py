from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import pandas as pd

def append_summary_to_existing_excel(file_path, summary_data, results, sim_results=None, top_10_hits=None):
    # Load workbook
    wb = load_workbook(file_path)
    if 'Summary' in wb.sheetnames:
        wb.remove(wb['Summary'])
    ws = wb.create_sheet(title='Summary')

    # Define styles
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_font = Font(bold=True)
    currency_format = '"$"#,##0.00'
    percent_format = '0.00%'

    # --- Table 1: Summary Data ---
    ws.append(['Summary Metric', 'Value'])
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        cell.fill = PatternFill(start_color="D9D9D9", fill_type="solid")

    for metric, value in summary_data.items():
        ws.append([metric, value])
        row = ws[ws.max_row]

        is_number = isinstance(value, (int, float))
        if is_number and metric == 'roi_percent':
            row[1].value = value / 100
            row[1].number_format = percent_format
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

    # Auto-size columns for first table
    for col in ['A', 'B']:
        ws.column_dimensions[col].auto_size = True

    # --- Table 1: Results Data ---
    next_row = ws.max_row + 2
    ws.cell(row=next_row, column=1, value='Results Metric')
    ws.cell(row=next_row, column=2, value='Value')
    for cell in ws[next_row]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        cell.fill = PatternFill(start_color="D9D9D9", fill_type="solid")

    for metric, value in results.items():
        next_row += 1
        ws.append([metric, value])
        row = ws[ws.max_row]
        is_number = isinstance(value, (int, float))
        if is_number and metric == 'roi_percent':
            row[1].value = value / 100
            row[1].number_format = percent_format
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

    # Auto-size columns for second table
    for col in ['A', 'B']:
        ws.column_dimensions[col].auto_size = True

    # --- Table 2: Simulation Stats & Percentiles ---
    next_row = ws.max_row + 2
    ws.cell(row=next_row, column=1, value='Simulation Metric')
    ws.cell(row=next_row, column=2, value='Value')
    for cell in ws[next_row]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        cell.fill = PatternFill(start_color="D9D9D9", fill_type="solid")

    sim_stats = [
        ('Mean Value', sim_results['mean']),
        ('Standard Deviation', sim_results['std_dev']),
        ('Minimum Value', sim_results['min']),
        ('Maximum Value', sim_results['max']),
    ]
    for metric, value in sim_stats:
        next_row += 1
        ws.append([metric, value])

    for perc_label, perc_val in sim_results['percentiles'].items():
        next_row += 1
        ws.append([perc_label, perc_val])

    # --- Table 3: Top 10 Most Expensive Hits ---
    next_row = ws.max_row + 2
    ws.cell(row=next_row, column=1, value='Top 10 Most Expensive Hits')
    ws.cell(row=next_row, column=2, value='Value')
    ws.cell(row=next_row, column=3, value='Effective Pull Rate')
    for cell in ws[next_row]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        cell.fill = PatternFill(start_color="D9D9D9", fill_type="solid")

    if top_10_hits is not None:
        if hasattr(top_10_hits, "values"):
            top_10_hits_rows = top_10_hits.values.tolist()
        else:
            top_10_hits_rows = top_10_hits
        for row_data in top_10_hits_rows:
            next_row += 1
            ws.append(row_data[:3])

    for col in ['A', 'B', 'C']:
        ws.column_dimensions[col].auto_size = True


    wb.save(file_path)
