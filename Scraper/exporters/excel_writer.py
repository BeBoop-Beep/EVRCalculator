from openpyxl import load_workbook
from openpyxl.styles import Alignment
from ..services.price_cleaner import clean_price_value

def save_to_excel(cards, prices, excel_path):
    """Save card and price data to Excel"""
    wb = load_workbook(excel_path)
    sheet = wb.active

    # Read current headers
    headers = [cell.value.strip() if cell.value else "" for cell in sheet[1]]
    existing_cols = {name: idx + 1 for idx, name in enumerate(headers)}

    desired_columns = [
        'Card Name', 'Rarity', 'Pull Rate (1/X)', 'Price ($)',
        'Reverse Variant Price ($)', 'Pack Price', 'Mini Tin Price',
        'Booster Bundle Price', 'ETB Price', 'ETB Promo Card Price',
        'Booster Box Price', 'Special Collection Price'
    ]

    column_indices = {}
    next_col = len(headers) + 1
    for col in desired_columns:
        if col in existing_cols:
            column_indices[col] = existing_cols[col]
        else:
            column_indices[col] = next_col
            sheet.cell(row=1, column=next_col, value=col)
            next_col += 1

    # Write card data
    for i, card in enumerate(cards, 2):
        sheet.cell(row=i, column=column_indices['Card Name'], value=card.get('productName', ''))
        sheet.cell(row=i, column=column_indices['Price ($)'], value=card.get('Price ($)', ''))
        sheet.cell(row=i, column=column_indices['Reverse Variant Price ($)'], value=card.get('Reverse Variant Price ($)', ''))
        sheet.cell(row=i, column=column_indices['Rarity'], value=card.get('rarity', ''))
        sheet.cell(row=i, column=column_indices['Pull Rate (1/X)'], value=card.get('Pull Rate (1/X)', ''))

        for col_idx in column_indices.values():
            sheet.cell(row=i, column=col_idx).alignment = Alignment(horizontal='left', vertical='center')

    # Write sealed prices
    price_mapping = {
        'Pack Price': 'Pack Price',
        'Mini Tin Price': 'Mini Tin Price',
        'Booster Bundle Price': 'Booster Bundle Price',
        'ETB Price': 'ETB Price',
        'ETB Promo Price': 'ETB Promo Card Price',
        'Booster Box Price': 'Booster Box Price',
        'Special Collection Price': 'Special Collection Price'
    }

    sealed_row = 2
    for price_type, price_value in prices.items():
        for key, col_name in price_mapping.items():
            if key.lower() in price_type.lower():
                col_idx = column_indices.get(col_name)
                if col_idx:
                    cleaned_price = clean_price_value(price_value)
                    sheet.cell(row=sealed_row, column=col_idx, value=cleaned_price if cleaned_price else "N/A")
                break

    # Auto-size columns
    for col_idx in column_indices.values():
        col_letter = sheet.cell(row=1, column=col_idx).column_letter
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in sheet[col_letter])
        sheet.column_dimensions[col_letter].width = max(15, max_length * 1.2)
    
    wb.save(excel_path)
    print(f"Successfully updated {len(cards)} cards and {len(prices)} sealed product prices.")