import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def determine_pull_rate(card_name, rarity_text, PULL_RATE_MAPPING):
    card_name_lower = card_name.lower()
    rarity_lower = rarity_text.lower().strip()

    # Check special cases first
    if 'master ball pattern' in card_name_lower:
        return PULL_RATE_MAPPING.get('master ball pattern'), 'master ball pattern'
    if 'poke ball pattern' in card_name_lower:
        return PULL_RATE_MAPPING.get('poke ball pattern'), 'poke ball pattern'
    if 'ace spec rare' in rarity_lower:
        return PULL_RATE_MAPPING.get('ace spec rare'), 'ace spec rare'

    # Use exact matching first
    for rarity_key in PULL_RATE_MAPPING:
        if rarity_lower == rarity_key:
            return PULL_RATE_MAPPING[rarity_key], rarity_key

    # Fallback fuzzy match in case of formatting issues
    for rarity_key in PULL_RATE_MAPPING:
        if rarity_key in rarity_lower:
            return PULL_RATE_MAPPING[rarity_key], rarity_key

    return None, rarity_text  # fallback to original rarity


# Function to fetch and parse card data
def fetch_price_data(price_guide_url, PULL_RATE_MAPPING):
    
    response = requests.get(price_guide_url)

    if response.status_code != 200:
        print(f"Error fetching data: {response.status_code}")
        print(response.text[:300])  # Show first part of the HTML/error for debugging
        return []

    try:
        json_data = response.json()
    except ValueError as e:
        print(f"Failed to parse JSON: {e}")
        print("Response content:", response.text[:300])  # Preview the bad content
        return []
    
    cards = json_data.get("result", [])

    card_data = {}

    for card in cards:
        product_name = card.get('productName')
        condition = card.get('condition')
        printing = card.get('printing')
        market_price = card.get('marketPrice')
        rarity = card.get('rarity')

        if not (product_name and condition and market_price and "Near Mint" in condition):
            continue
 
        if "code card" in product_name.lower():
            continue

        is_reverse = 'Reverse' in printing or 'Reverse Holofoil' in printing

        if product_name not in card_data:
            pull_rate, normalized_rarity = determine_pull_rate(product_name, rarity, PULL_RATE_MAPPING)

            card_data[product_name] = {
                'productName': product_name,
                'Price ($)': '',
                'Reverse Variant Price ($)': '',
                'rarity': normalized_rarity,
                'Pull Rate (1/X)': pull_rate
            }

        if is_reverse:
            card_data[product_name]['Reverse Variant Price ($)'] = market_price
        else:
            card_data[product_name]['Price ($)'] = market_price

    # Convert dict to list for Excel saving
    return list(card_data.values())

def clean_price_value(price_str):
    if not price_str:
        return None
    if isinstance(price_str, (int, float)):
        return price_str
    price_str = price_str.strip()
    if price_str.lower().startswith("$"):
        try:
            return float(price_str[1:])
        except ValueError:
            return None
    if price_str.lower() in ["unavailable", "n/a", "no url"]:
        return None
    try:
        # Try converting directly to float if no $
        return float(price_str)
    except ValueError:
        return None


def fetch_product_market_price(price_url):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/113.0.0.0 Safari/537.36",
        "Accept": "application/json",
        "Referer": "https://www.tcgplayer.com/"
    }
     
    response = requests.get(price_url, headers=headers)
    if response.status_code != 200:
        print(f"Failed to fetch data from {price_url}: {response.status_code}")
        return None

    try:
        data = response.json()
        first_market_price = data.get("result", [])[0].get("buckets", [])[0].get("marketPrice", None)
        return first_market_price
    except (IndexError, AttributeError, ValueError) as e:
        print(f"Error parsing data from {price_url}: {e}")
        return None

# Loop through all endpoints and print prices
def get_all_first_market_prices(endpoints):
    price_results = {}

    for label, url in endpoints.items():
        if not url:
            print(f"{label}: $Unavailable (No URL)")
            continue

        price = fetch_product_market_price(url)
        if price is not None:
            print(f"{label}: ${price}")
            price_results[label] = price
        else:
            print(f"{label}: $Unavailable")

    return price_results


# Function to save data to Excel
def save_to_excel(cards, prices, excel_path):
    wb = load_workbook(excel_path)
    sheet = wb.active

    # Read current headers (row 1)
    headers = [cell.value.strip() if cell.value else "" for cell in sheet[1]]
    existing_cols = {name: idx + 1 for idx, name in enumerate(headers)}

    # Desired columns in logical order
    desired_columns = [
        'Card Name',
        'Rarity',
        'Pull Rate (1/X)',
        'Price ($)',
        'Reverse Variant Price ($)',
        'Pack Price',
        'Mini Tin Price',
        'Booster Bundle Price',
        'ETB Price',
        'ETB Promo Card Price',
        'Booster Box Price',
        'Special Collection Price'
    ]

    # Build safe column indices: append missing ones to the end
    column_indices = {}
    next_col = len(headers) + 1
    for col in desired_columns:
        if col in existing_cols:
            column_indices[col] = existing_cols[col]
        else:
            column_indices[col] = next_col
            sheet.cell(row=1, column=next_col, value=col)
            next_col += 1

    # Write card data
    for i, card in enumerate(cards, 2):
        sheet.cell(row=i, column=column_indices['Card Name'], value=card.get('productName', ''))
        sheet.cell(row=i, column=column_indices['Price ($)'], value=card.get('Price ($)', ''))
        sheet.cell(row=i, column=column_indices['Reverse Variant Price ($)'], value=card.get('Reverse Variant Price ($)', ''))
        sheet.cell(row=i, column=column_indices['Rarity'], value=card.get('rarity', ''))
        sheet.cell(row=i, column=column_indices['Pull Rate (1/X)'], value=card.get('Pull Rate (1/X)', ''))

        # Align formatting
        for col_idx in column_indices.values():
            sheet.cell(row=i, column=col_idx).alignment = Alignment(horizontal='left', vertical='center')


    # Mapping for matching price keys to column names
    price_mapping = {
        'Pack Price': 'Pack Price',
        'Mini Tin Price': 'Mini Tin Price',
        'Booster Bundle Price': 'Booster Bundle Price',
        'ETB Price': 'ETB Price',
        'ETB Promo Price': 'ETB Promo Card Price',
        'Booster Box Price': 'Booster Box Price',
        'Special Collection Price': 'Special Collection Price'
    }

    sealed_row = 2  # fixed row for sealed prices

    for price_type, price_value in prices.items():
        # find matching column and write price on row 2
        for key, col_name in price_mapping.items():
            if key.lower() in price_type.lower():
                col_idx = column_indices.get(col_name)
                if col_idx:
                    cleaned_price = clean_price_value(price_value)
                    if cleaned_price is not None:
                        sheet.cell(row=sealed_row, column=col_idx, value=cleaned_price)
                    else:
                        sheet.cell(row=sealed_row, column=col_idx, value="N/A")
                break

    print(f"DEBUG: Cards to save: {len(cards)}")

    # Optional: adjust column widths
    for col_idx in column_indices.values():
        col_letter = sheet.cell(row=1, column=col_idx).column_letter
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in sheet[col_letter])
        sheet.column_dimensions[col_letter].width = max(15, max_length * 1.2)
    
    wb.save(excel_path)
    print(f"Successfully updated {len(cards)} cards and {len(prices)} sealed product prices.")
    # print(f"DEBUG: Current headers: {headers}")
    # print(f"DEBUG: Column indices: {column_indices}")
    # print(f"DEBUG: Sealed prices: {prices}")
    # print(f"DEBUG: Sealed prices row: {sealed_row}")



# Main function
def scrape_tcgplayer_xhr(excel_path="ev_output.xlsx", config={}):
    cards = fetch_price_data(config.SCRAPE_URL, config.PULL_RATE_MAPPING)

    prices = get_all_first_market_prices(config.PRICE_ENDPOINTS)
    
    save_to_excel(cards, prices, excel_path)

if __name__ == "__main__":
    scrape_tcgplayer_xhr()
