from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import pandas as pd

def append_summary_to_existing_excel(file_path, summary_data, results):
    # Merge and prepare data
    combined_data = {**summary_data, **results}

    # Fix percent fields that are already in percentage format
    for key in ['roi_percent', 'no_hit_probability_percentage', 'hit_probability_percentage']:
        if key in combined_data:
            combined_data[key] = combined_data[key]

    df = pd.DataFrame({
        "Metric": list(combined_data.keys()),
        "Value": list(combined_data.values())
    })

    # Load workbook
    wb = load_workbook(file_path)

    # Remove old Summary sheet if it exists
    if 'Summary' in wb.sheetnames:
        wb.remove(wb['Summary'])

    # Create new Summary sheet
    ws = wb.create_sheet(title='Summary')

    # Define styles
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_font = Font(bold=True)
    currency_format = '"$"#,##0.00'
    percent_format = '0.00%'

    # Write headers
    ws.append(['Metric', 'Value'])
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        cell.fill = PatternFill(start_color="D9D9D9", fill_type="solid")

    # Write rows
    for metric, value in combined_data.items():
        ws.append([metric, value])
        row = ws[ws.max_row]

        # Ensure value is a number before applying number formats
        is_number = isinstance(value, (int, float))
        dollar_keys = [
            'ev_common_total', 'ev_uncommon_total', 'ev_rare_total', 'ev_double_rare_total',
            'ev_pokeball_total', 'ev_master_ball_total', 'ev_hyper_rare_total',
            'ev_ultra_rare_total', 'ev_SIR_total', 'ev_IR_total', 'ev_reverse_total',
            'ev_total_for_hits', 'ev_hits_total', 'total_ev', 'net_value', 'ev_ace_spec_total'
        ]

        if is_number and metric in dollar_keys:
            row[1].number_format = currency_format
        elif is_number and metric in ['roi_percent', 'hit_probability_percentage', 'no_hit_probability_percentage']:
            row[1].value = value / 100  # Convert to Excel percent
            row[1].number_format = percent_format
        elif is_number and metric == 'roi':
            row[1].number_format = '0.00'


        # Borders & alignment
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

    # Auto-size columns
    for col in ['A', 'B']:
        ws.column_dimensions[col].auto_size = True

    # Save workbook
    wb.save(file_path)
