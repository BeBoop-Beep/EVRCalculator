import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def determine_pull_rate(card_name, rarity_text, PULL_RATE_MAPPING):
    card_name_lower = card_name.lower()
    rarity_lower = rarity_text.lower().strip()

    # Check special cases first
    if 'poke ball pattern' in card_name_lower:
        return 302
    if 'master ball pattern' in card_name_lower:
        return 1362
    if 'ace spec' in rarity_lower:
        return 128

    # Use exact matching first
    for rarity_key in PULL_RATE_MAPPING:
        if rarity_lower == rarity_key:
            return PULL_RATE_MAPPING[rarity_key]

    # Fallback fuzzy match in case of formatting issues
    for rarity_key in PULL_RATE_MAPPING:
        if rarity_key in rarity_lower:
            return PULL_RATE_MAPPING[rarity_key]

    return None


# Function to fetch and parse card data
def fetch_price_data(price_guide_url, PULL_RATE_MAPPING):
    response = requests.get(price_guide_url)
    
    if response.status_code != 200:
        print(f"Error fetching data: {response.status_code}")
        return []

    json_data = response.json()
    cards = json_data.get("result", [])

    card_data = {}

    for card in cards:
        product_name = card.get('productName')
        condition = card.get('condition')
        printing = card.get('printing')
        market_price = card.get('marketPrice')
        rarity = card.get('rarity')

        if not (product_name and condition and market_price and "Near Mint" in condition):
            continue
 
        if "code card" in product_name.lower():
            continue

        is_reverse = 'Reverse' in printing or 'Reverse Holofoil' in printing

        if product_name not in card_data:
            pull_rate = determine_pull_rate(product_name, rarity, PULL_RATE_MAPPING)

            card_data[product_name] = {
                'productName': product_name,
                'Price ($)': '',
                'Reverse Variant Price ($)': '',
                'rarity': rarity,
                'Pull Rate (1/X)': pull_rate
            }

        if is_reverse:
            card_data[product_name]['Reverse Variant Price ($)'] = market_price
        else:
            card_data[product_name]['Price ($)'] = market_price

    # Convert dict to list for Excel saving
    return list(card_data.values())

# Function to save data to Excel
def save_to_excel(cards, excel_path):
    wb = load_workbook(excel_path)
    sheet = wb.active

    # Find or create columns
    headers = [cell.value for cell in sheet[1]]
    column_indices = {
        'Card Name': headers.index('Card Name') + 1 if 'Card Name' in headers else len(headers) + 1,
        'Price ($)': headers.index('Price ($)') + 1 if 'Price ($)' in headers else len(headers) + 2,
        'Reverse Variant Price ($)': headers.index('Reverse Variant Price ($)') + 1 if 'Reverse Variant Price ($)' in headers else len(headers) + 3,
        'Rarity': headers.index('Rarity') + 1 if 'Rarity' in headers else len(headers) + 4,
        'Pull Rate (1/X)': headers.index('Pull Rate (1/X)') + 1 if 'Pull Rate (1/X)' in headers else len(headers) + 5,
    }


    # Update headers if needed
    for col_name, col_idx in column_indices.items():
        if sheet.cell(row=1, column=col_idx).value != col_name:
            sheet.cell(row=1, column=col_idx, value=col_name)

    # Write data
    for i, card in enumerate(cards, 2):
        sheet.cell(row=i, column=column_indices['Card Name'], value=card['productName'])
        sheet.cell(row=i, column=column_indices['Price ($)'], value=card.get('Price ($)', ''))
        sheet.cell(row=i, column=column_indices['Reverse Variant Price ($)'], value=card.get('Reverse Variant Price ($)', ''))
        sheet.cell(row=i, column=column_indices['Rarity'], value=card.get('rarity', ''))
        sheet.cell(row=i, column=column_indices['Pull Rate (1/X)'], value=card.get('Pull Rate (1/X)', ''))

        # Align formatting
        for col_idx in column_indices.values():
            sheet.cell(row=i, column=col_idx).alignment = Alignment(horizontal='left', vertical='center')

    # Adjust column widths
    for col_letter in ['A', 'B', 'C', 'D', 'E']:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in sheet[col_letter])
        sheet.column_dimensions[col_letter].width = max(15, max_length) * 1.2

    wb.save(excel_path)
    print(f"Successfully updated {len(cards)} cards.")

# Main function
def scrape_tcgplayer_xhr(excel_path="ev_output.xlsx", price_guide_url=None, PULL_RATE_MAPPING={}):
    price_guide_url = 'https://infinite-api.tcgplayer.com/priceguide/set/23237/cards/?rows=5000&productTypeID=1'
    cards = fetch_price_data(price_guide_url, PULL_RATE_MAPPING)
    save_to_excel(cards, excel_path)

if __name__ == "__main__":
    scrape_tcgplayer_xhr()
