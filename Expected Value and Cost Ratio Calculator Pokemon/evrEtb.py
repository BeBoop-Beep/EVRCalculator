import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os

# Load the spreadsheet
file_path = "excelDocs/prismaticEvolution/pokemon_data_final.xlsx"
df = pd.read_excel(file_path, engine='openpyxl')

# Ensure minimum required columns exist
required_columns = ["Current ETB Market Price", "Current Market Price of ETB Promo Card", "Total EV"]
for col in required_columns:
    if col not in df.columns:
        raise KeyError(f"Missing required column: '{col}' in the spreadsheet.")

# Market price per etb (update accordingly)
etb_price = df["Current ETB Market Price"].iloc[0]

# Market price per etb (update accordingly)
etb_promo_price = df["Current Market Price of ETB Promo Card"].iloc[0]

total_ev_per_pack = df['Total EV'].iloc[0]

# Pack configuration constants
TOTAL_PACKS_PER_ETB = 9 

# Calculate base EV before hit rate adjustment
expected_profit_in_etb = (total_ev_per_pack * TOTAL_PACKS_PER_ETB) + etb_promo_price
Etb_Net_Value = expected_profit_in_etb - etb_price
Etb_Roi = expected_profit_in_etb / etb_price

# Add all calculations to the DataFrame for inspection
df.loc[0, "Total ETB EV"] = expected_profit_in_etb
df.loc[0, "ETB NET VALUE"] = Etb_Net_Value
df.loc[0, "ETB ROI"] = Etb_Roi

# Save to new file
base_filename = os.path.splitext(file_path)[0]
output_file = base_filename + "_with_etb.xlsx"

# Preserve formatting
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name="Data")
    workbook = writer.book
    worksheet = workbook.active
    
    # Copy column widths from original
    original_workbook = load_workbook(file_path)
    original_worksheet = original_workbook.active
    for i, col in enumerate(original_worksheet.columns):
        max_length = max(len(str(cell.value)) for cell in col) if any(cell.value for cell in col) else 0
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[get_column_letter(i + 1)].width = adjusted_width

# Print detailed EV breakdown
print("\nHit Rate Calculation:")
print(f"Total EV_ETB: {expected_profit_in_etb}")
print(f"Total Etb_Net_Value: {Etb_Net_Value}")
print(f"Total Etb_Roi: {Etb_Roi}")
