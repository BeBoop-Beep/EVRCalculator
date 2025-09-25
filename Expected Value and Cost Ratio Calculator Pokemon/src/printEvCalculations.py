from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import pandas as pd

def append_summary_to_existing_excel(file_path, summary_data, results, sim_results=None, top_10_hits=None):
    # Merge and prepare data
    combined_data = {**summary_data, **results}

    # Fix percent fields that are already in percentage format
    for key in ['roi_percent', 'no_hit_probability_percentage', 'hit_probability_percentage']:
        if key in combined_data:
            combined_data[key] = combined_data[key]

    df = pd.DataFrame({
        "Metric": list(combined_data.keys()),
        "Value": list(combined_data.values())
    })

    # Load workbook
    wb = load_workbook(file_path)
    if 'Summary' in wb.sheetnames:
        wb.remove(wb['Summary'])
    ws = wb.create_sheet(title='Summary')

    # Define styles
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_font = Font(bold=True)
    currency_format = '"$"#,##0.00'
    percent_format = '0.00%'

    # Write headers
    ws.append(['Metric', 'Value'])
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        cell.fill = PatternFill(start_color="D9D9D9", fill_type="solid")

    # Write rows
    for metric, value in combined_data.items():
        ws.append([metric, value])
        row = ws[ws.max_row]

        # Ensure value is a number before applying number formats
        is_number = isinstance(value, (int, float))
        dollar_keys = [
            'ev_common_total', 'ev_uncommon_total', 'ev_rare_total', 'ev_double_rare_total',
            'ev_pokeball_total', 'ev_master_ball_total', 'ev_hyper_rare_total',
            'ev_ultra_rare_total', 'ev_SIR_total', 'ev_IR_total', 'ev_reverse_total',
            'ev_total_for_hits', 'ev_hits_total', 'total_ev', 'net_value', 'ev_ace_spec_total'
        ]

        if is_number and metric in dollar_keys:
            row[1].number_format = currency_format
        elif is_number and metric in ['roi_percent', 'hit_probability_percentage', 'no_hit_probability_percentage']:
            row[1].value = value / 100  # Convert to Excel percent
            row[1].number_format = percent_format
        elif is_number and metric == 'roi':
            row[1].number_format = '0.00'


        # Borders & alignment
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

    # Auto-size columns
    for col in ['A', 'B']:
        ws.column_dimensions[col].auto_size = True

    # Table 2: Simulation Stats & Percentiles
    next_row = ws.max_row + 2
    ws.cell(row=next_row, column=1, value='Simulation Metric')
    ws.cell(row=next_row, column=2, value='Value')
    for cell in ws[next_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        cell.fill = PatternFill(start_color="D9D9D9", fill_type="solid")

    sim_stats = [
        ('Mean Value', sim_results['mean']),
        ('Standard Deviation', sim_results['std_dev']),
        ('Minimum Value', sim_results['min']),
        ('Maximum Value', sim_results['max']),
    ]
    for metric, value in sim_stats:
        next_row += 1
        ws.append([metric, value])

    # Percentiles
    for perc_label, perc_val in sim_results['percentiles'].items():
        next_row += 1
        ws.append([perc_label, perc_val])

    # Table 3: Top 10 Most Expensive Hits
    next_row = ws.max_row + 2
    ws.cell(row=next_row, column=1, value='Top 10 Most Expensive Hits')
    ws.cell(row=next_row, column=2, value='Value')
    ws.cell(row=next_row, column=3, value='Effective Pull Rate')
    for cell in ws[next_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        cell.fill = PatternFill(start_color="D9D9D9", fill_type="solid")

    # If top_10_hits is a DataFrame
    if top_10_hits is not None:
        # If it's a DataFrame, convert to list of lists
        if hasattr(top_10_hits, "values"):
            top_10_hits_rows = top_10_hits.values.tolist()
        else:
            top_10_hits_rows = top_10_hits  # fallback for list of tuples

        for row_data in top_10_hits_rows:
            next_row += 1
            ws.append(row_data[:3])  # Only take first 3 columns if more are present

    # Auto-size columns
    for col in ['A', 'B', 'C']:
        ws.column_dimensions[col].auto_size = True

    wb.save(file_path)
