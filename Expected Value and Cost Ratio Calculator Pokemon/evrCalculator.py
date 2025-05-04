import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os

def load_rarity_config(wb):
    """Load rarity configuration from '_RarityConfig' sheet"""
    try:
        config_sheet = wb['_RarityConfig']
        config = {}
        for row in config_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:  # If both columns have values
                config[str(row[0]).strip().lower()] = float(row[1])
        return config
    except KeyError:
        print("No '_RarityConfig' sheet found - using default values")
        return {}

def calculate_pack_ev(file_path):
    """
    Calculate pack expected value (EV) and related metrics from a Pokémon card spreadsheet.
    """
    # Load workbook first to get rarity config
    wb = load_workbook(file_path)
    rarity_config = load_rarity_config(wb)
    
    # Now load data with pandas
    df = pd.read_excel(file_path, engine='openpyxl')

    # Ensure minimum required columns exist
    required_columns = ["Card Name", "Pull Rate (1/X)", "Price ($)", "Hit Rate Adjustment (1+HR)"]
    for col in required_columns:
        if col not in df.columns:
            raise KeyError(f"Missing required column: '{col}' in the spreadsheet.")

    # Market price per pack
    pack_price = df["Current Market Pack Price"].iloc[0]

    # Dynamic hit rarities based on config
    HIT_RARITIES = [
        "master ball pattern",
        "special illustration rare",
        "illustration rare",
        "ace spec",
        "hyper rare",
        "ultra rare"
    ] if rarity_config is None else [
        k for k in rarity_config.keys() 
        if any(x in k for x in ['rare', 'spec', 'hyper', 'ultra', 'illustration'])
    ]

    # Function to identify hits
    def is_hit(row):
        card_name = str(row["Card Name"]).lower()
        rarity = str(row["Rarity"]).lower() if "Rarity" in df.columns else ""
        if "master ball pattern" in card_name:
            return True
        return any(hit_rarity in rarity for hit_rarity in HIT_RARITIES)

    # Calculate hit rate adjustment
    total_cards = len(df)
    total_hits = sum(df.apply(is_hit, axis=1))
    hit_rate_adjustment = 1 + (total_hits / total_cards)
    df.loc[df.index[0], "Hit Rate Adjustment (1+HR)"] = hit_rate_adjustment

    # Define rarity groups and multipliers (can be moved to config if needed)
    RARITY_GROUPS = {
        "common": 5.5,
        "uncommon": 1.5,
        "rare": 1.5
    }

    def classify_card(row, rarity_config):
        if rarity_config is None:
                rarity_config = {}
        """Classify cards with 0 as default for unconfigured rarities"""
        card_name = str(row["Card Name"]).lower()
        rarity = str(row["Rarity"]).lower() if "Rarity" in row.index else ""
        
        # 1. Check config first (returns 0 if rarity not found)
        for config_rarity in sorted(rarity_config.keys(), key=len, reverse=True):
            if config_rarity in rarity or config_rarity in card_name:
                return (
                    rarity_config.get(config_rarity, 0),  # Default 0 if key exists but rate is missing
                    "common" if "common" in config_rarity else
                    "uncommon" if "uncommon" in config_rarity else
                    "rare" if "rare" in config_rarity else "other"
                )
        
        # 2. Hardcoded fallbacks with 0 defaults
        if "poke ball pattern" in card_name:
            return (rarity_config.get('poke ball pattern', 0), "other")
        if "master ball pattern" in card_name:
            return (rarity_config.get('master ball pattern', 0), "other")
        if "ace spec" in rarity:
            return (rarity_config.get('ace spec', 0), "other")
        
        # 3. Base rarities
        if rarity == "common":
            return (rarity_config.get('common', 0), "common")
        if rarity == "uncommon":
            return (rarity_config.get('uncommon', 0), "uncommon")
        if rarity in ["rare", "holo rare"]:
            return (rarity_config.get('rare', 0), "rare")
        
        # 4. Everything else gets 0
        return (0, "other")

    # Classify all cards
    df[["Pull Rate", "Rarity Group"]] = df.apply(
    lambda row: pd.Series(classify_card(row, rarity_config=rarity_config)), axis=1)
    
    # [Rest of your original calculate_pack_ev function...]
    # ... including EV calculations, saving, and results preparation

  
    # Calculate EV components for each group
    ev_components = {
        "common": df[df["Rarity Group"] == "common"]["Price ($)"].div(
                  df[df["Rarity Group"] == "common"]["Pull Rate"]).sum(),
        "uncommon": df[df["Rarity Group"] == "uncommon"]["Price ($)"].div(
                    df[df["Rarity Group"] == "uncommon"]["Pull Rate"]).sum(),
        "rare": df[df["Rarity Group"] == "rare"]["Price ($)"].div(
                df[df["Rarity Group"] == "rare"]["Pull Rate"]).sum(),
        "other": df[df["Rarity Group"] == "other"]["Price ($)"].div(
                 df[df["Rarity Group"] == "other"]["Pull Rate"]).sum()
    }

    # Apply multipliers and calculate totals
    common_ev = ev_components["common"] * RARITY_GROUPS["common"]
    uncommon_ev = ev_components["uncommon"] * RARITY_GROUPS["uncommon"]
    rare_ev = ev_components["rare"] * RARITY_GROUPS["rare"]
    other_ev = ev_components["other"] * 1  # multiplier of 1

    # Calculate base EV before hit rate adjustment
    EV_base = common_ev + uncommon_ev + rare_ev + other_ev
    EV = EV_base * hit_rate_adjustment
    net_value = EV - pack_price
    roi_per_pack = EV / pack_price

    # Add all calculations to the DataFrame for inspection
    df["EV Component"] = df["Price ($)"] / df["Pull Rate"]
    df.loc[0, "Total Cards"] = total_cards
    df.loc[0, "Total Hits"] = total_hits
    df.loc[0, "Hit Rate Adjustment"] = hit_rate_adjustment
    df.loc[0, "Common EV Sum"] = ev_components["common"]
    df.loc[0, "Uncommon EV Sum"] = ev_components["uncommon"]
    df.loc[0, "Rare EV Sum"] = ev_components["rare"]
    df.loc[0, "Other EV Sum"] = ev_components["other"]
    df.loc[0, "Common EV (x5.5)"] = common_ev
    df.loc[0, "Uncommon EV (x1.5)"] = uncommon_ev
    df.loc[0, "Rare EV (x1.5)"] = rare_ev
    df.loc[0, "Other EV"] = other_ev
    df.loc[0, "Total EV Base"] = EV_base
    df.loc[0, "Total EV"] = EV
    df.loc[0, "net value per pack"] = net_value
    df.loc[0, "roi per pack"] = roi_per_pack

    # Save to new file
    base_filename = os.path.splitext(file_path)[0]
    output_file = base_filename + "_final.xlsx"

    # Preserve formatting
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Data")
        workbook = writer.book
        worksheet = workbook.active
        
        # Copy column widths from original
        original_workbook = load_workbook(file_path)
        original_worksheet = original_workbook.active
        for i, col in enumerate(original_worksheet.columns):
            max_length = max(len(str(cell.value)) for cell in col) if any(cell.value for cell in col) else 0
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[get_column_letter(i + 1)].width = adjusted_width

    # Prepare results dictionary
    results = {
        "total_cards": total_cards,
        "total_hits": total_hits,
        "hit_rate_adjustment": hit_rate_adjustment,
        "ev_components": ev_components,
        "common_ev": common_ev,
        "uncommon_ev": uncommon_ev,
        "rare_ev": rare_ev,
        "other_ev": other_ev,
        "EV_base": EV_base,
        "EV": EV,
        "net_value": net_value,
        "roi_per_pack": roi_per_pack,
        "output_file": output_file,
        "use_rarity_adjustment": use_rarity_adjustment
    }

    return results, output_file

def print_results(results):
    """Print the calculation results in a readable format"""
    print("\nHit Rate Calculation:")
    print(f"Total Cards: {results['total_cards']}")
    print(f"Total Hits: {results['total_hits']}")
    print(f"Hit Rate Adjustment: 1 + ({results['total_hits']}/{results['total_cards']}) = {results['hit_rate_adjustment']:.4f}")

    print("\nDetailed EV Calculation Breakdown:")
    print(f"Sum of (Price/PullRate) for Commons: {results['ev_components']['common']:.6f} x5.5 = {results['common_ev']:.6f}")
    print(f"Sum of (Price/PullRate) for Uncommons: {results['ev_components']['uncommon']:.6f} x1.5 = {results['uncommon_ev']:.6f}")
    print(f"Sum of (Price/PullRate) for Rares: {results['ev_components']['rare']:.6f} x1.5 = {results['rare_ev']:.6f}")
    print(f"Sum of (Price/PullRate) for Others: {results['ev_components']['other']:.6f} x1 = {results['other_ev']:.6f}")
    print(f"\nBase EV (before hit rate): {results['EV_base']:.6f}")
    print(f"Final EV per pack: ${results['EV']:.2f}")
    print(f"net_value: {results['net_value']:.4f}")
    print(f"roi_per_pack: {results['roi_per_pack']:.4f}")

    print(f"\nUpdated spreadsheet saved as {results['output_file']}!")

    if not results['use_rarity_adjustment']:
        print("Note: 'Rarity' column not found. Used default 1/X pull rates.")
