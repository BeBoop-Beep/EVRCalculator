from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import os

def load_rarity_config(wb):
    """Load rarity configuration from '_RarityConfig' sheet in the workbook"""
    try:
        config_sheet = wb['_RarityConfig']
        config = {}
        for row in config_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:  # If we have both rarity name and pull rate
                config[str(row[0]).strip().lower()] = float(row[1])
        return config
    except KeyError:
        # Fallback to default values if config sheet doesn't exist
        return {}

def determine_pull_rate(card_name, rarity_text, rarity_mapping):
    """Determine pull rate based on card name and rarity text (optimized)."""
    card_name_lower = card_name.lower()
    rarity_lower = rarity_text.lower()

    # Check special cases first (exact name matches)
    if 'poke ball pattern' in card_name_lower:
        return rarity_mapping.get('poke ball pattern', 302)
    if 'master ball pattern' in card_name_lower:
        return rarity_mapping.get('master ball pattern', 1362)
    if 'ace spec' in rarity_lower:  # Moved here since it's a special case
        return rarity_mapping.get('ace spec', 128)

    # Define priority checks (order matters for overlapping terms)
    rarity_checks = [
        ('special illustration rare', 'special illustration rare'),
        ('hyper rare', 'hyper rare'),
        ('master ball pattern', 'master ball pattern'),  # Already handled, but kept for clarity
        ('poke ball pattern', 'poke ball pattern'),      # Already handled, but kept for clarity
        ('double rare', 'double rare'),
        ('ultra rare', 'ultra rare'),
        ('rare', ' rare '),  # Space ensures we don't match "double rare" or "ultra rare"
        ('uncommon', 'uncommon'),
        ('common', ' common '),  # Space prevents matching "uncommon"
    ]

    for (rarity_key, match_term) in rarity_checks:
        if match_term in f' {rarity_lower} ':  # Add spaces to avoid partial matches
            return rarity_mapping.get(rarity_key)

    return None  # No match found

def parse_card_data(html_content, rarity_mapping):
    soup = BeautifulSoup(html_content, 'html.parser')
    cards = []
    
    rows = soup.select('tbody.tcg-table-body tr')
    
    for row in rows:
        try:
            # 1. EXTRACT CARD NAME (unchanged)
            name = "Unknown"
            name_element = row.select_one('a.pdp-url')
            if name_element:
                name = name_element.get_text(strip=True).split('<svg')[0].strip()
                if 'Code Card' in name:
                    continue
            
            # 2. EXTRACT PRICE (unchanged)
            price = "Price not found"
            price_cells = row.select('td.tcg-table-body__cell--align-right')
            for cell in price_cells:
                text = cell.get_text(strip=True)
                if text.startswith('$'):
                    price = text
                    break
            
            # 3. NEW PRECISE RARITY EXTRACTION
            rarity = "Unknown"
            # Look for the specific cell pattern that contains rarity
            rarity_cell = row.select_one('td.tcg-table-body__cell--align-left:-soup-contains("Illustration")'
                                      ', td.tcg-table-body__cell--align-left:-soup-contains("Rare")'
                                      ', td.tcg-table-body__cell--align-left:-soup-contains("Uncommon")'
                                      ', td.tcg-table-body__cell--align-left:-soup-contains("Common")')
            
            if rarity_cell:
                rarity = rarity_cell.get_text(strip=True)
            
            # 4. SPECIAL CASE FOR ACE SPEC (not in name but in rarity)
            if 'ACE SPEC' in rarity:
                pull_rate = rarity_mapping.get('ace spec', 128)
            else:
                pull_rate = determine_pull_rate(name, rarity, rarity_mapping)
            
            cards.append({
                'Card Name': name,
                'Price ($)': price,
                'Rarity': rarity,
                'Pull Rate (1/X)': pull_rate
            })
            
        except Exception as e:
            print(f"Error processing row: {e}")
            continue
    
    return cards

def htmlScraper(excel_path):
    with open("page_content.html", "r", encoding="utf-8") as f:
        html_content = f.read()

    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found at {excel_path}")

    # Load workbook and rarity config first
    wb = load_workbook(excel_path)
    rarity_mapping = load_rarity_config(wb)
    
    # Parse card data with the loaded rarity mapping
    cards = parse_card_data(html_content, rarity_mapping)

    sheet = wb.active

    # Find or create columns (now includes 'Rarity')
    headers = [cell.value for cell in sheet[1]]
    column_indices = {
        'Card Name': headers.index('Card Name') + 1 if 'Card Name' in headers else len(headers) + 1,
        'Price ($)': headers.index('Price ($)') + 1 if 'Price ($)' in headers else len(headers) + 2,
        'Rarity': headers.index('Rarity') + 1 if 'Rarity' in headers else len(headers) + 3,
        'Pull Rate (1/X)': headers.index('Pull Rate (1/X)') + 1 if 'Pull Rate (1/X)' in headers else len(headers) + 4
    }

    # Update headers if needed
    for col_name, col_idx in column_indices.items():
        if sheet.cell(row=1, column=col_idx).value != col_name:
            sheet.cell(row=1, column=col_idx, value=col_name)

    # Write data (now includes 'Rarity')
    for i, card in enumerate(cards, 2):  # Start from row 2
        sheet.cell(row=i, column=column_indices['Card Name'], value=card['Card Name'])
        
        # Handle price conversion carefully
        price_str = card['Price ($)']
        if price_str.startswith('$'):
            try:
                price_value = float(price_str.replace('$', '').replace(',', ''))
            except ValueError:
                price_value = price_str  # Keep as string if conversion fails
        else:
            price_value = price_str
        
        sheet.cell(row=i, column=column_indices['Price ($)'], value=price_value)
        sheet.cell(row=i, column=column_indices['Rarity'], value=card['Rarity'])
        sheet.cell(row=i, column=column_indices['Pull Rate (1/X)'], value=card['Pull Rate (1/X)'])
        
        # Apply formatting
        for col_idx in column_indices.values():
            sheet.cell(row=i, column=col_idx).alignment = Alignment(horizontal='left', vertical='center')

    # Adjust column widths (update if needed)
    for col_letter in ['A', 'B', 'C', 'D']:  # Now includes column D (Rarity)
        max_length = max(
            len(str(cell.value)) if cell.value else 0
            for cell in sheet[col_letter]
        )
        sheet.column_dimensions[col_letter].width = max(15, max_length) * 1.2

    wb.save(excel_path)
    print(f"Successfully updated {len(cards)} cards with pull rates and rarities")

if __name__ == "__main__":
    htmlScraper()